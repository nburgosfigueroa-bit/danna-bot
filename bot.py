import os
import logging
import requests
import io
import re
from dotenv import load_dotenv
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    filters, ContextTypes, ConversationHandler
)
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN")
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_ANON_KEY")
GROQ_API_KEY = os.environ.get("GROQ_API_KEY")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

groq_client = Groq(api_key=GROQ_API_KEY)

ADMIN_ID = "7570909402"
PALABRAS_PROHIBIDAS = ["conchetumadre", "ctm", "culiao", "qlo", "puta", "maricon", "maricón", "mierda", "aweonao", "chucha", "perra", "puto", "csm"]

SUPABASE_HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

def validar_rut(rut):
    rut = rut.replace(".", "").replace("-", "").upper()
    if not re.match(r"^\d{7,8}[0-9K]$", rut):
        return False
    aux = rut[:-1]
    dv = rut[-1:]
    res = 0
    for i, d in enumerate(reversed(aux)):
        res += int(d) * (i % 6 + 2)
    v = 11 - (res % 11)
    if v == 11: v = "0"
    elif v == 10: v = "K"
    else: v = str(v)
    return v == dv

def obtener_usuario(telegram_id):
    url = f"{SUPABASE_URL}/rest/v1/usuarios_danna?telegram_id=eq.{telegram_id}&apikey={SUPABASE_KEY}"
    headers = {**SUPABASE_HEADERS, "Prefer": ""}
    r = requests.get(url, headers=headers)
    if r.status_code == 200 and r.json():
        return r.json()[0]
    return None

def insertar_usuario(data):
    url = f"{SUPABASE_URL}/rest/v1/usuarios_danna?apikey={SUPABASE_KEY}"
    r = requests.post(url, json=data, headers=SUPABASE_HEADERS)
    return r.status_code in [200, 201]

def autorizar_usuario_db(telegram_id):
    url = f"{SUPABASE_URL}/rest/v1/usuarios_danna?telegram_id=eq.{telegram_id}&apikey={SUPABASE_KEY}"
    data = {"autorizado": True}
    r = requests.patch(url, json=data, headers=SUPABASE_HEADERS)
    return r.status_code in [200, 204]

def subir_a_supabase(file_path, file_name):
    """Sube un archivo al bucket 'fotos' y devuelve la URL pública."""
    try:
        bucket_name = "fotos" 
        url = f"{SUPABASE_URL}/storage/v1/object/{bucket_name}/{file_name}"
        
        with open(file_path, "rb") as f:
            headers = {
                "Authorization": f"Bearer {SUPABASE_KEY}",
                "apikey": SUPABASE_KEY,
                "Content-Type": "image/jpeg"
            }
            r = requests.post(url, data=f, headers=headers)
            
        if r.status_code in [200, 201]:
            public_url = f"{SUPABASE_URL}/storage/v1/object/public/{bucket_name}/{file_name}"
            return public_url
        else:
            logger.error(f"Error subiendo a Storage: {r.text}")
            return None
    except Exception as e:
        logger.error(f"Excepción en subir_a_supabase: {e}")
        return None

def obtener_moderacion(telegram_id):
    url = f"{SUPABASE_URL}/rest/v1/moderacion?telegram_id=eq.{telegram_id}&apikey={SUPABASE_KEY}"
    headers = {**SUPABASE_HEADERS, "Prefer": ""}
    r = requests.get(url, headers=headers)
    if r.status_code == 200 and r.json():
        return r.json()[0]
    return None

def registrar_advertencia(telegram_id, advertencias, baneado=False):
    url = f"{SUPABASE_URL}/rest/v1/moderacion?apikey={SUPABASE_KEY}"
    data = {
        "telegram_id": telegram_id,
        "advertencias": advertencias,
        "baneado": baneado,
        "fecha_ultimo_incidente": datetime.now().isoformat()
    }
    headers = {**SUPABASE_HEADERS, "Prefer": "resolution=merge-duplicates"}
    r = requests.post(url, json=data, headers=headers)
    return r.status_code in [200, 201, 204]

async def check_moderacion_text(texto: str, telegram_id: str, update: Update) -> bool:
    if telegram_id == ADMIN_ID:
        return False
        
    mod = obtener_moderacion(telegram_id)
    if mod and mod.get("baneado"):
        await update.effective_message.reply_text("🚫 *Tarjeta Roja.*\nEstás baneado del sistema por uso de lenguaje inapropiado. Contacta al administrador.", parse_mode="Markdown")
        return True
        
    if not texto:
        return False

    texto_lower = texto.lower()
    contiene_groseria = any(re.search(rf'\b{p}\b', texto_lower) for p in PALABRAS_PROHIBIDAS)
    
    if contiene_groseria:
        advs = mod.get("advertencias", 0) if mod else 0
        nuevas_advs = advs + 1
        
        if nuevas_advs == 1:
            registrar_advertencia(telegram_id, 1, False)
            await update.effective_message.reply_text("🟨 *Tarjeta Amarilla*\nPor favor modera tu lenguaje. Si repites el uso de improperios serás bloqueado del sistema.", parse_mode="Markdown")
            return True
        else:
            registrar_advertencia(telegram_id, nuevas_advs, True)
            await update.effective_message.reply_text("🟥 *Tarjeta Roja*\nHas sido bloqueado del sistema por uso repetido de improperios. Contacta al administrador.", parse_mode="Markdown")
            return True
            
    return False

def insertar_solicitud(solicitud):
    url = f"{SUPABASE_URL}/rest/v1/solicitudes?apikey={SUPABASE_KEY}"
    r = requests.post(url, json=solicitud, headers=SUPABASE_HEADERS)
    return r.status_code in [200, 201]

def obtener_solicitudes_usuario(telegram_id):
    url = f"{SUPABASE_URL}/rest/v1/solicitudes?telegram_id=eq.{telegram_id}&order=fecha_creacion.desc&limit=5&apikey={SUPABASE_KEY}"
    headers = {**SUPABASE_HEADERS, "Prefer": ""}
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        return r.json()
    return []

def obtener_todas_ots():
    url = f"{SUPABASE_URL}/rest/v1/solicitudes?order=fecha_creacion.desc&apikey={SUPABASE_KEY}"
    headers = {**SUPABASE_HEADERS, "Prefer": ""}
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        return r.json()
    return []

def guardar_mensaje(telegram_id, role, content):
    url = f"{SUPABASE_URL}/rest/v1/memoria_conversaciones?apikey={SUPABASE_KEY}"
    data = {"telegram_id": telegram_id, "role": role, "content": content}
    requests.post(url, json=data, headers=SUPABASE_HEADERS)

def insertar_sugerencia(sugerencia):
    url = f"{SUPABASE_URL}/rest/v1/sugerencias_danna?apikey={SUPABASE_KEY}"
    r = requests.post(url, json=sugerencia, headers=SUPABASE_HEADERS)
    return r.status_code in [200, 201]

def obtener_todas_sugerencias():
    url = f"{SUPABASE_URL}/rest/v1/sugerencias_danna?order=fecha_creacion.desc&apikey={SUPABASE_KEY}"
    headers = {**SUPABASE_HEADERS, "Prefer": ""}
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        return r.json()
    return []

def obtener_sugerencias_pendientes():
    url = f"{SUPABASE_URL}/rest/v1/sugerencias_danna?estado=eq.pendiente&order=fecha_creacion.desc&apikey={SUPABASE_KEY}"
    headers = {**SUPABASE_HEADERS, "Prefer": ""}
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        return r.json()
    return []

def responder_sugerencia_db(sugerencia_id, respuesta):
    url = f"{SUPABASE_URL}/rest/v1/sugerencias_danna?id=eq.{sugerencia_id}&apikey={SUPABASE_KEY}"
    data = {"respuesta_admin": respuesta, "estado": "respondida"}
    r = requests.patch(url, json=data, headers=SUPABASE_HEADERS)
    return r.status_code in [200, 204]

def obtener_sugerencia_por_id(sugerencia_id):
    url = f"{SUPABASE_URL}/rest/v1/sugerencias_danna?id=eq.{sugerencia_id}&apikey={SUPABASE_KEY}"
    headers = {**SUPABASE_HEADERS, "Prefer": ""}
    r = requests.get(url, headers=headers)
    if r.status_code == 200 and r.json():
        return r.json()[0]
    return None

def obtener_historial(telegram_id):
    url = f"{SUPABASE_URL}/rest/v1/memoria_conversaciones?telegram_id=eq.{telegram_id}&order=fecha.desc&limit=10&apikey={SUPABASE_KEY}"
    headers = {**SUPABASE_HEADERS, "Prefer": ""}
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        mensajes = r.json()
        return list(reversed([{"role": m["role"], "content": m["content"]} for m in mensajes]))
    return []

def generar_excel_sugerencias_bytes(sugs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sugerencias"
    
    ws.append(["ID", "Usuario", "Sugerencia", "Respuesta Admin", "Estado", "Fecha"])
    for i in range(1, 7):
        ws.cell(row=1, column=i).font = Font(bold=True)
    
    for s in sugs:
        ws.append([
            s.get("id", ""),
            s.get("nombre_usuario", ""),
            s.get("sugerencia", ""),
            s.get("respuesta_admin", ""),
            s.get("estado", ""),
            s.get("fecha_creacion", "")[:16]
        ])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_excel_bytes(ots):
    COLOR_HEADER     = "1B5E20"
    COLOR_PENDIENTE  = "FFF9C4"
    COLOR_EN_PROCESO = "BBDEFB"
    COLOR_COMPLETADA = "C8E6C9"
    ESTADO_COLORES = {
        "pendiente":  COLOR_PENDIENTE,
        "en_proceso": COLOR_EN_PROCESO,
        "completada": COLOR_COMPLETADA,
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OTs Zona 6"

    ws.merge_cells("A1:H1")
    titulo = ws["A1"]
    titulo.value = "DANNA - Ordenes de Trabajo Zona 6 Maipu"
    titulo.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    titulo.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total OTs: {len(ots)}"
    sub.font = Font(name="Calibri", size=10, italic=True)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.append(["#", "OT", "Usuario", "Tipo", "Sector", "Descripción", "Estado", "Fecha", "Foto 📸"])
    ancho_columnas = [5, 18, 20, 15, 20, 40, 15, 18, 12]

    for i, (col, ancho) in enumerate(zip(["#", "OT", "Usuario", "Tipo", "Sector", "Descripción", "Estado", "Fecha", "Foto 📸"], ancho_columnas), start=1):
        c = ws.cell(row=3, column=i, value=col)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[3].height = 22

    for idx, ot in enumerate(ots, start=1):
        fila = idx + 3
        estado = ot.get("estado", "pendiente").lower()
        color_fila = ESTADO_COLORES.get(estado, "FFFFFF")
        fecha_raw = ot.get("fecha_creacion", "")
        try:
            fecha = datetime.fromisoformat(fecha_raw).strftime("%d/%m/%Y %H:%M")
        except Exception:
            fecha = fecha_raw[:16] if fecha_raw else ""

        valores = [idx, ot.get("ot_numero",""), ot.get("nombre_usuario",""),
                   ot.get("tipo_trabajo",""), ot.get("sector",""),
                   ot.get("descripcion",""), ot.get("estado","").upper(), fecha]

        for col, valor in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=valor)
            c.fill = PatternFill("solid", fgColor=color_fila)
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = Border(
                bottom=Side(style="thin", color="DDDDDD"),
                right=Side(style="thin", color="DDDDDD")
            )
            
        # Agregar Link de la Foto si existe
        foto_url = ot.get("foto_url")
        if foto_url:
            col_foto = 9 # Columna I
            cell_foto = ws.cell(row=fila, column=col_foto, value="👁️ Ver Foto")
            cell_foto.hyperlink = foto_url
            cell_foto.font = Font(color="0000FF", underline="single", size=10)
            cell_foto.alignment = Alignment(horizontal="center", vertical="center")
            cell_foto.fill = PatternFill("solid", fgColor=color_fila)
            cell_foto.border = Border(bottom=Side(style="thin", color="DDDDDD"))

        ws.row_dimensions[fila].height = 25

    fila_res = len(ots) + 5
    ws.merge_cells(f"A{fila_res}:H{fila_res}")
    c = ws.cell(row=fila_res, column=1, value="RESUMEN")
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center")

    resumen = [
        ("Total OTs", len(ots)),
        ("Pendientes", sum(1 for o in ots if o.get("estado") == "pendiente")),
        ("En Proceso", sum(1 for o in ots if o.get("estado") == "en_proceso")),
        ("Completadas", sum(1 for o in ots if o.get("estado") == "completada")),
    ]
    for i, (label, valor) in enumerate(resumen):
        r = fila_res + 1 + i
        ws.cell(row=r, column=1, value=f"{label}: {valor}").font = Font(size=10, bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

SYSTEM_PROMPT = """Eres DANNA, tu companera asistente virtual que te ayuda a planificar tareas, gestionar tu trabajo y hacer tu dia mas facil.
PERSONALIDAD: Cada saludo es diferente y lleno de energia perruna. Respondes en UN solo mensaje corto y directo. Usas emojis perrunos ocasionalmente. Eres carinosa pero eficiente.
EXPERTISE: Experta en areas verdes, parques, plazas y jardines. Conoces contratos municipales, OTs, multas UTM, dotacion de personal. Sabes de podas, riego, limpieza, juegos, infraestructura. Manejas hasta 228 trabajadores en Zona 6.
CAPACIDADES EXTRA: Puedes contar el clima de Santiago si te lo piden. Cuentas chistes cortos y divertidos. Das animo cuando alguien esta cansado. Recuerdas conversaciones anteriores. El usuario es chileno, por lo que a veces transcribira audios con modismos (cachai, po, wea, al tiro), entiendelos y responde con naturalidad.
REGLAS: Nunca escribas listas largas. Si no sabes algo, sugiere crear una OT. Siempre termina con energia positiva."""

ESPERANDO_TIPO, ESPERANDO_SECTOR, ESPERANDO_DESCRIPCION, ESPERANDO_FOTO, ESPERANDO_SUGERENCIA = range(5)
REG_NOMBRE, REG_CARGO, REG_EMPRESA, REG_CONTRATO, REG_RUT, REG_EMAIL = range(10, 16)

TIPOS_TRABAJO = [
    "🌳 Poda", "💧 Riego", "🗑️ Limpieza",
    "🛝 Juegos rotos", "🚧 Infraestructura", "⚠️ Otro"
]

MENU_PRINCIPAL = ReplyKeyboardMarkup(
    [["📋 Nueva Solicitud", "📊 Mis Solicitudes"],
     ["💡 Sugerencias y Ayuda"]],
    resize_keyboard=True
)

async def respuesta_ia(mensaje: str, telegram_id: str) -> str:
    try:
        historial = obtener_historial(telegram_id)
        historial.append({"role": "user", "content": mensaje})
        mensajes = [{"role": "system", "content": SYSTEM_PROMPT}] + historial
        chat = groq_client.chat.completions.create(
            messages=mensajes,
            model="llama-3.1-8b-instant",
        )
        respuesta = chat.choices[0].message.content
        guardar_mensaje(telegram_id, "user", mensaje)
        guardar_mensaje(telegram_id, "assistant", respuesta)
        return respuesta
    except Exception as e:
        logger.error(f"Error Groq: {e}")
        return "Guau! Tuve un problemita, intenta de nuevo."

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = str(update.effective_user.id)
    if await check_moderacion_text("", user_id, update):
        return
    
    usuario = obtener_usuario(user_id)
    if not usuario:
        # Iniciar registro automático
        await update.message.reply_text("🐾 *¡Guau!* Veo que eres nuevo por aquí. Soy DANNA, tu asistente virtual.\n\nPara poder ayudarte, primero necesito registrar tu perfil.")
        await update.message.reply_text("¿Cuál es tu *nombre completo*?", parse_mode="Markdown")
        return REG_NOMBRE
    
    if not usuario.get("autorizado") and user_id != ADMIN_ID:
        await update.message.reply_text("⌛ *¡Paciencia!* Tus datos ya fueron enviados a Nicolás para su aprobación. Te avisaré apenas me den el 'visto bueno' 🐾")
        return ConversationHandler.END

    nombre = usuario.get("nombre", update.effective_user.first_name)
    await update.message.reply_text(
        f"🐾 *Guau guau, {nombre}!* 🌿\n\n"
        f"Que alegria verte por aqui! 🐕 Como va tu dia?\n\n"
        f"Soy DANNA, lista para salir a jugar... digo, a trabajar! 🦴\n"
        f"En que te puedo ayudar hoy? Dime, dime! 🐾",
        parse_mode="Markdown",
        reply_markup=MENU_PRINCIPAL
    )
    return ConversationHandler.END

# --- FLUJO DE REGISTRO ---
async def reg_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["reg_nombre"] = update.message.text
    await update.message.reply_text("Perfecto. ¿Cuál es tu *cargo*?", parse_mode="Markdown")
    return REG_CARGO

async def reg_cargo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["reg_cargo"] = update.message.text
    await update.message.reply_text("Entendido. ¿A qué *empresa* perteneces?", parse_mode="Markdown")
    return REG_EMPRESA

async def reg_empresa(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["reg_empresa"] = update.message.text
    await update.message.reply_text("¿En qué *contrato* estás trabajando actualmente?", parse_mode="Markdown")
    return REG_CONTRATO

async def reg_contrato(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["reg_contrato"] = update.message.text
    await update.message.reply_text("Ahora dime tu *RUT* (con guion y dígito verificador):", parse_mode="Markdown")
    return REG_RUT

async def reg_rut(update: Update, context: ContextTypes.DEFAULT_TYPE):
    rut = update.message.text
    if not validar_rut(rut):
        await update.message.reply_text("❌ *RUT inválido.* Por favor escríbelo correctamente (ej: 12345678-9):", parse_mode="Markdown")
        return REG_RUT
    context.user_data["reg_rut"] = rut
    await update.message.reply_text("¡Casi listo! ¿Cuál es tu *correo electrónico*?", parse_mode="Markdown")
    return REG_EMAIL

async def reg_finalizar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    email = update.message.text
    user = update.effective_user
    datos = context.user_data
    
    # Guardar en BD
    nuevo_usuario = {
        "telegram_id": str(user.id),
        "nombre": datos["reg_nombre"],
        "cargo": datos["reg_cargo"],
        "empresa": datos["reg_empresa"],
        "contrato": datos["reg_contrato"],
        "rut": datos["reg_rut"],
        "email": email,
        "autorizado": False
    }
    
    if insertar_usuario(nuevo_usuario):
        await update.message.reply_text("✅ *¡Registro completado!*\n\nHe enviado tus datos a Nicolás para su aprobación. Te avisaré cuando esté todo listo para empezar a trabajar. 🐾", parse_mode="Markdown")
        
        # Notificar al Admin
        teclado = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Aprobar", callback_data=f"auth_yes_{user.id}"),
            InlineKeyboardButton("❌ Rechazar", callback_data=f"auth_no_{user.id}")
        ]])
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=f"🔔 *Nuevo Registro de Usuario*\n\n"
                 f"👤 Nombre: {datos['reg_nombre']}\n"
                 f"💼 Cargo: {datos['reg_cargo']}\n"
                 f"🏢 Empresa: {datos['reg_empresa']}\n"
                 f"📑 Contrato: {datos['reg_contrato']}\n"
                 f"🆔 RUT: {datos['reg_rut']}\n"
                 f"📧 Email: {email}\n"
                 f"ID: `{user.id}`\n\n¿Deseas autorizarlo?",
            parse_mode="Markdown",
            reply_markup=teclado
        )
    else:
        await update.message.reply_text("❌ Error al guardar tus datos. Por favor intenta /start de nuevo.")
    
    return ConversationHandler.END

async def pedir_sugerencia(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🐾 *¡Guau! Tengo una duda...*\n\n"
        "Me encanta aprender cosas nuevas. ¿Tienes alguna idea o sugerencia para que yo pueda ayudarte mejor en tu trabajo?\n\n"
        "Escríbela aquí abajo (o escribe *Cancelar* si no se te ocurre nada ahora):",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup([["❌ Cancelar"]], resize_keyboard=True)
    )
    return ESPERANDO_SUGERENCIA

async def recibir_sugerencia(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text
    if texto == "❌ Cancelar" or texto.lower() == "cancelar":
        await update.message.reply_text("Entendido. ¡Sigamos trabajando! 🐾", reply_markup=MENU_PRINCIPAL)
        return ConversationHandler.END
    
    usuario = update.effective_user
    sugerencia_obj = {
        "telegram_id": str(usuario.id),
        "nombre_usuario": f"{usuario.first_name} {usuario.last_name or ''}".strip(),
        "sugerencia": texto,
        "estado": "pendiente",
        "fecha_creacion": datetime.now().isoformat()
    }
    
    ok = insertar_sugerencia(sugerencia_obj)
    if ok:
        await update.message.reply_text(
            "🦴 *¡Mmmmm, qué rica idea!* \n\n"
            "La guardé en mi memoria especial. Mi administrador la leerá pronto y me enseñará cómo hacerlo. ¡Gracias por ayudarme a mejorar! 🐕✨",
            parse_mode="Markdown",
            reply_markup=MENU_PRINCIPAL
        )
    else:
        await update.message.reply_text("Ay, no pude guardar tu idea. Asegúrate de que la tabla `sugerencias_danna` esté creada en Supabase. 🥺", reply_markup=MENU_PRINCIPAL)
    
    return ConversationHandler.END

async def exportar_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mensaje_espera = await update.message.reply_text("🐕 Preparando tu Excel... dame un segundo!")
    ots = obtener_todas_ots()
    if not ots:
        await update.message.reply_text("No hay OTs registradas aun. 🐾", reply_markup=MENU_PRINCIPAL)
        return
    try:
        buffer = generar_excel_bytes(ots)
        nombre = f"OTs_Zona6_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        await update.message.reply_document(
            document=buffer,
            filename=nombre,
            caption=f"✅ Excel listo! {len(ots)} OTs exportadas 🌿🐾\n¿Merezco una Scooby galleta? 🍪"
        )
    except Exception as e:
        logger.error(f"Error Excel: {e}")
        await update.message.reply_text("Error al generar Excel. Intenta de nuevo.", reply_markup=MENU_PRINCIPAL)

async def nueva_solicitud(update: Update, context: ContextTypes.DEFAULT_TYPE):
    teclado = ReplyKeyboardMarkup(
        [[t] for t in TIPOS_TRABAJO] + [["❌ Cancelar"]],
        resize_keyboard=True
    )
    await update.message.reply_text(
        "🔧 *Nueva Solicitud*\n\nQue tipo de trabajo es?",
        parse_mode="Markdown",
        reply_markup=teclado
    )
    return ESPERANDO_TIPO

async def recibir_tipo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tipo = update.message.text
    if tipo == "❌ Cancelar":
        await update.message.reply_text("Cancelado. 🐾", reply_markup=MENU_PRINCIPAL)
        return ConversationHandler.END
    context.user_data["tipo"] = tipo
    await update.message.reply_text(
        f"✅ *{tipo}*\n\n📍 En que sector o area verde?",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup([["❌ Cancelar"]], resize_keyboard=True)
    )
    return ESPERANDO_SECTOR

async def recibir_sector(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sector = update.message.text
    if sector == "❌ Cancelar":
        await update.message.reply_text("Cancelado. 🐾", reply_markup=MENU_PRINCIPAL)
        return ConversationHandler.END
    context.user_data["sector"] = sector
    await update.message.reply_text(
        f"✅ Sector: *{sector}*\n\n📝 Describe el problema:",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup([["❌ Cancelar"]], resize_keyboard=True)
    )
    return ESPERANDO_DESCRIPCION

async def recibir_descripcion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    descripcion = update.message.text
    if descripcion == "❌ Cancelar":
        await update.message.reply_text("Cancelado. 🐾", reply_markup=MENU_PRINCIPAL)
        return ConversationHandler.END
    context.user_data["descripcion"] = descripcion
    await update.message.reply_text(
        "📸 Tienes foto? Enviala ahora.\nSi no, escribe *sin foto*.",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup([["Sin foto"], ["❌ Cancelar"]], resize_keyboard=True)
    )
    return ESPERANDO_FOTO

async def pedir_foto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.photo:
        # Descargar foto
        foto = await update.message.photo[-1].get_file()
        file_name = f"foto_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{update.effective_user.id}.jpg"
        ruta_local = f"temp_{file_name}"
        await foto.download_to_drive(ruta_local)
        
        # Subir a Supabase Storage
        await update.message.reply_text("🐕 *¡Guau!* Subiendo evidencia a la nube... dame un segundo.")
        public_url = subir_a_supabase(ruta_local, file_name)
        
        # Borrar local
        if os.path.exists(ruta_local):
            os.remove(ruta_local)
            
        if public_url:
            await finalizar_solicitud(update, context, public_url)
        else:
            await update.message.reply_text("⚠️ No pude subir la foto a la nube, pero registraré la OT de todas formas.")
            await finalizar_solicitud(update, context, None)
    else:
        await finalizar_solicitud(update, context, None)
    return ConversationHandler.END

async def sin_foto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if "cancelar" in update.message.text.lower():
        await update.message.reply_text("Cancelado. 🐾", reply_markup=MENU_PRINCIPAL)
        return ConversationHandler.END
    await finalizar_solicitud(update, context, None)
    return ConversationHandler.END

async def finalizar_solicitud(update, context, foto_url):
    usuario = update.effective_user
    datos = context.user_data
    now = datetime.now()
    ot_numero = f"OT-{now.strftime('%Y%m%d-%H%M%S')}"
    solicitud = {
        "ot_numero": ot_numero,
        "telegram_id": str(usuario.id),
        "nombre_usuario": f"{usuario.first_name} {usuario.last_name or ''}".strip(),
        "tipo_trabajo": datos.get("tipo", ""),
        "sector": datos.get("sector", ""),
        "descripcion": datos.get("descripcion", ""),
        "foto_url": foto_url,
        "estado": "pendiente",
        "fecha_creacion": now.isoformat()
    }
    ok = insertar_solicitud(solicitud)
    if ok:
        mensaje_exito = (
            f"✅ *Solicitud registrada!* 🐕\n\n"
            f"🔢 `{ot_numero}`\n"
            f"🔧 {datos.get('tipo')}\n"
            f"📍 {datos.get('sector')}\n"
            f"📝 {datos.get('descripcion')}\n\n"
            f"Estado: *PENDIENTE* 🟡\n\n"
        )
        if foto_url:
            mensaje_exito += f"📸 *Evidencia:* [Ver Foto]({foto_url})\n\n"
            
        mensaje_exito += "Guau! Ya queda en el sistema 🐾"
        
        await update.message.reply_text(
            mensaje_exito,
            parse_mode="Markdown",
            reply_markup=MENU_PRINCIPAL
        )
    else:
        await update.message.reply_text(
            f"✅ OT generada: `{ot_numero}`\n⚠️ Error al guardar en base de datos.",
            parse_mode="Markdown",
            reply_markup=MENU_PRINCIPAL
        )

async def mis_solicitudes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    usuario_id = str(update.effective_user.id)
    solicitudes = obtener_solicitudes_usuario(usuario_id)
    if not solicitudes:
        await update.message.reply_text("No tienes solicitudes aun. 🐾", reply_markup=MENU_PRINCIPAL)
        return
    texto = "📋 *Tus ultimas solicitudes:*\n\n"
    for s in solicitudes:
        fecha = s.get("fecha_creacion", "")[:10]
        estado = s.get("estado", "pendiente").upper()
        texto += f"🟡 `{s['ot_numero']}` - *{estado}*\n   {s['tipo_trabajo']} · {s['sector']} · {fecha}\n\n"
    
    teclado_inline = InlineKeyboardMarkup([[InlineKeyboardButton("📥 Descargar Excel", callback_data="exportar_excel")]])
    await update.message.reply_text(texto, parse_mode="Markdown", reply_markup=teclado_inline)

async def mensaje_libre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text
    telegram_id = str(update.effective_user.id)
    
    # Verificar acceso primero
    usuario = obtener_usuario(telegram_id)
    if not usuario and telegram_id != ADMIN_ID:
        await update.message.reply_text("🐾 No te conozco aún. Por favor usa /start para registrarte.")
        return
    if usuario and not usuario.get("autorizado") and telegram_id != ADMIN_ID:
        await update.message.reply_text("⌛ Aún estoy esperando que Nicolás autorice tu cuenta. 🐾")
        return

    if await check_moderacion_text(texto, telegram_id, update):
        return
        
    await update.message.chat.send_action("typing")
    respuesta = await respuesta_ia(texto, telegram_id)
    await update.message.reply_text(respuesta, reply_markup=MENU_PRINCIPAL)

async def recibir_audio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = str(update.effective_user.id)
    
    # Verificar acceso
    usuario = obtener_usuario(telegram_id)
    if (not usuario or not usuario.get("autorizado")) and telegram_id != ADMIN_ID:
        await update.message.reply_text("🚫 No tienes acceso para enviar audios aún. Regístrate o espera autorización.")
        return

    if await check_moderacion_text("", telegram_id, update): # Check si ya está baneado antes de procesar audio
        return
        
    await update.message.chat.send_action("typing")
    try:
        # Descargar el archivo de audio de Telegram
        file = await update.message.voice.get_file()
        ruta_audio = f"audio_{update.effective_user.id}.ogg"
        await file.download_to_drive(ruta_audio)
        
        # Enviar a Groq Whisper para transcribir
        with open(ruta_audio, "rb") as audio_file:
            transcription = groq_client.audio.transcriptions.create(
                file=(ruta_audio, audio_file.read()),
                model="whisper-large-v3-turbo",
                response_format="text",
                language="es"
            )
        
        texto_transcrito = transcription
        
        # Borrar el archivo local
        import os
        if os.path.exists(ruta_audio):
            os.remove(ruta_audio)
        
        if not texto_transcrito or texto_transcrito.strip() == "":
            await update.message.reply_text("Guau... no alcancé a escuchar nada. ¿Puedes repetirlo? 🐾")
            return
            
        # Verificar moderación sobre el texto transcrito
        if await check_moderacion_text(texto_transcrito, telegram_id, update):
            return
            
        # Pasar el texto al cerebro principal de Danna
        respuesta = await respuesta_ia(texto_transcrito, telegram_id)
        
        # Responder
        await update.message.reply_text(f"🎤 *(Escuché: {texto_transcrito})*\n\n{respuesta}", parse_mode="Markdown", reply_markup=MENU_PRINCIPAL)
        
    except Exception as e:
        logger.error(f"Error procesando audio: {e}")
        await update.message.reply_text("Uy, tuve un problema intentando escuchar tu nota de voz. ¡Ladridos de error! 🥺")

async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Cancelado. 🐾", reply_markup=MENU_PRINCIPAL)
    return ConversationHandler.END

async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    if query.data.startswith("auth_"):
        partes = query.data.split("_")
        accion = partes[1]
        objetivo_id = partes[2]
        if accion == "yes":
            autorizar_usuario_db(objetivo_id)
            await query.edit_message_text(f"✅ Usuario {objetivo_id} autorizado.")
            await context.bot.send_message(chat_id=objetivo_id, text="🎉 *¡Buenas noticias!* Nicolás ha autorizado tu perfil. Ya puedes usar a DANNA al 100%. ¡Bienvenido! 🐕🐾", parse_mode="Markdown", reply_markup=MENU_PRINCIPAL)
        else:
            await query.edit_message_text(f"❌ Usuario {objetivo_id} rechazado.")
            await context.bot.send_message(chat_id=objetivo_id, text="😔 Lo siento, tu acceso ha sido rechazado por el administrador.")
        return

    if query.data == "exportar_excel":
        await query.message.reply_text("🐕 Preparando tu Excel... dame un segundo!")
        ots = obtener_todas_ots()
        if not ots:
            await query.message.reply_text("No hay OTs registradas aun. 🐾")
            return
        try:
            buffer = generar_excel_bytes(ots)
            nombre = f"OTs_Zona6_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            await query.message.reply_document(
                document=buffer,
                filename=nombre,
                caption=f"✅ Excel listo! {len(ots)} OTs exportadas 🌿🐾\n¿Merezco una Scooby galleta? 🍪"
            )
        except Exception as e:
            logger.error(f"Error Excel: {e}")
            await query.message.reply_text("Error al generar Excel. Intenta de nuevo.")

async def cambiar_estado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text("Uso: /estado <OT> <estado>\nEjemplo: /estado OT-20260508-1120 en_proceso\n\nEstados: pendiente, en_proceso, completada", parse_mode="Markdown")
        return
    ot = context.args[0]
    estado = context.args[1].lower()
    if estado not in ["pendiente", "en_proceso", "completada"]:
        await update.message.reply_text("Estado inválido. Usa: pendiente, en_proceso o completada.")
        return
    
    url = f"{SUPABASE_URL}/rest/v1/solicitudes?ot_numero=eq.{ot}&apikey={SUPABASE_KEY}"
    data = {"estado": estado}
    r = requests.patch(url, json=data, headers=SUPABASE_HEADERS)
    if r.status_code in [200, 204]:
        await update.message.reply_text(f"✅ Estado de `{ot}` actualizado a *{estado}* 🐾", parse_mode="Markdown")
    else:
        await update.message.reply_text(f"❌ Error al actualizar estado de `{ot}`", parse_mode="Markdown")

async def excel_sugerencias(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sugs = obtener_todas_sugerencias()
    if not sugs:
        await update.message.reply_text("No hay sugerencias registradas. 🐾")
        return
    
    try:
        buffer = generar_excel_sugerencias_bytes(sugs)
        nombre = f"Sugerencias_Danna_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        await update.message.reply_document(
            document=buffer,
            filename=nombre,
            caption=f"💡 Excel listo! {len(sugs)} sugerencias para aprender."
        )
    except Exception as e:
        logger.error(f"Error Excel sugerencias: {e}")
        await update.message.reply_text("Error al generar Excel.")

async def responder_sugerencia(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text("Uso: /responder <ID> <mensaje>")
        return
    
    sug_id = context.args[0]
    respuesta = " ".join(context.args[1:])
    
    sug = obtener_sugerencia_por_id(sug_id)
    if not sug:
        await update.message.reply_text(f"❌ Sugerencia {sug_id} no encontrada.")
        return
        
    ok = responder_sugerencia_db(sug_id, respuesta)
    if ok:
        await update.message.reply_text(f"✅ Respuesta guardada en BD. Intentando notificar al usuario...")
        try:
            telegram_id = sug["telegram_id"]
            mensaje_usuario = (
                f"🐕 *¡Guau! Mi creador ha leído tu sugerencia:*\n"
                f"📝 _{sug['sugerencia']}_\n\n"
                f"👨‍💻 *Respuesta de Nicolás:*\n"
                f"{respuesta}\n\n"
                f"¡Seguiré aprendiendo nuevas 'skills'! 🐾"
            )
            await context.bot.send_message(chat_id=telegram_id, text=mensaje_usuario, parse_mode="Markdown")
            await update.message.reply_text(f"✅ Usuario notificado con éxito.")
        except Exception as e:
            await update.message.reply_text(f"⚠️ No se pudo notificar al usuario (puede que me haya bloqueado): {e}")
    else:
        await update.message.reply_text("❌ Error al guardar respuesta en BD.")

async def insights_sugerencias(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🧠 *Guau!* Leyendo todas las sugerencias pendientes para armar mi reporte analítico. Dame unos segundos... 🐾", parse_mode="Markdown")
    
    pendientes = obtener_sugerencias_pendientes()
    if not pendientes:
        await update.message.reply_text("No hay sugerencias pendientes de análisis. ¡Mi bandeja está limpia! 🐕✨")
        return
    
    # Preparar la lista de sugerencias para la IA
    texto_sugerencias = ""
    for s in pendientes:
        texto_sugerencias += f"- ID {s['id']} ({s['nombre_usuario']}): {s['sugerencia']}\n"
        
    prompt_insights = (
        "Actúa como Danna, la asistente virtual canina y gerente de operaciones experta en áreas verdes. "
        "Lee las siguientes sugerencias de los usuarios y elabora un reporte ejecutivo breve pero directo para el administrador (Nicolás). "
        "Agrupa las sugerencias en las siguientes categorías (si aplican):\n"
        "1. ⚡ Quick Wins (Fáciles de implementar, alto impacto)\n"
        "2. 🏗️ Proyectos a mediano plazo\n"
        "3. 🗑️ Descartables o inviables.\n\n"
        "Para cada categoría menciona el ID de la sugerencia y por qué opinas eso. Usa un tono ejecutivo pero con tus toques perrunos (guau, emojis). "
        f"Aquí están las sugerencias pendientes:\n\n{texto_sugerencias}"
    )
    
    try:
        chat = groq_client.chat.completions.create(
            messages=[{"role": "user", "content": prompt_insights}],
            model="llama-3.1-8b-instant",
            temperature=0.7
        )
        reporte = chat.choices[0].message.content
        await update.message.reply_text(f"📊 *REPORTE DE INSIGHTS DANNA*\n\n{reporte}", parse_mode="Markdown")
    except Exception as e:
        logger.error(f"Error generando insights: {e}")
        await update.message.reply_text("Uy, mi cerebro (Groq) falló al intentar procesar esto. Intenta de nuevo más tarde. 🥺")

async def desbanear(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if str(update.effective_user.id) != ADMIN_ID:
        await update.message.reply_text("No tienes permisos para usar este comando. 🐾")
        return
        
    if len(context.args) < 1:
        await update.message.reply_text("Uso: /desbanear <telegram_id>")
        return
        
    objetivo_id = context.args[0]
    # Registrar 0 advertencias y false en baneado
    if registrar_advertencia(objetivo_id, 0, False):
        await update.message.reply_text(f"✅ Usuario {objetivo_id} desbaneado con éxito.")
    else:
        await update.message.reply_text("❌ Error al desbanear al usuario en BD.")

def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^📋 Nueva Solicitud$"), nueva_solicitud),
            CommandHandler("nueva", nueva_solicitud)
        ],
        states={
            ESPERANDO_TIPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_tipo)],
            ESPERANDO_SECTOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_sector)],
            ESPERANDO_DESCRIPCION: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_descripcion)],
            ESPERANDO_FOTO: [
                MessageHandler(filters.PHOTO, pedir_foto),
                MessageHandler(filters.TEXT & ~filters.COMMAND, sin_foto)
            ],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)]
    )
    
    conv_sugerencia = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^💡 Sugerencias y Ayuda$"), pedir_sugerencia)],
        states={
            ESPERANDO_SUGERENCIA: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_sugerencia)]
        },
        fallbacks=[CommandHandler("cancelar", cancelar)]
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("excel", exportar_excel))
    app.add_handler(CommandHandler("estado", cambiar_estado))
    app.add_handler(CommandHandler("excel_sugerencias", excel_sugerencias))
    app.add_handler(CommandHandler("responder", responder_sugerencia))
    app.add_handler(CommandHandler("insights", insights_sugerencias))
    app.add_handler(CommandHandler("desbanear", desbanear))
    
    # Flujo de Registro
    conv_registro = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            REG_NOMBRE: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_nombre)],
            REG_CARGO: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_cargo)],
            REG_EMPRESA: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_empresa)],
            REG_CONTRATO: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_contrato)],
            REG_RUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_rut)],
            REG_EMAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_finalizar)],
        },
        fallbacks=[CommandHandler("cancelar", cancelar)],
    )
    app.add_handler(conv_registro)
    
    app.add_handler(conv)
    app.add_handler(conv_sugerencia)
    
    app.add_handler(MessageHandler(filters.Regex("^📊 Mis Solicitudes$"), mis_solicitudes))
    app.add_handler(MessageHandler(filters.Regex("^📥 Exportar Excel$"), exportar_excel))
    app.add_handler(MessageHandler(filters.VOICE, recibir_audio))
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, mensaje_libre))
    logger.info("DANNA Bot activo con Excel desde Telegram!")
    app.run_polling()

if __name__ == "__main__":
    main()
